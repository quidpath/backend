"""Utility helpers for the Excel financial extractor."""

from __future__ import annotations

import logging
import re
import unicodedata
from dataclasses import dataclass
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from difflib import SequenceMatcher
from typing import TYPE_CHECKING, Any, Dict, Iterable, Mapping, Optional, Sequence, Tuple

from .exceptions import ValueNormalizationError

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

NUMERIC_RE = re.compile(r"[^\d\.\-]")
CURRENCY_PREFIX_RE = re.compile(r"^(kes|ksh|kshs|ksh\.)\s*", flags=re.IGNORECASE)
WHITESPACE_RE = re.compile(r"\s+")


def normalize_label_text(label: str) -> str:
    """
    Normalize label text for fuzzy matching.

     FIX: Handles unicode hyphens, special characters, and whitespace.
    """
    if not label or not isinstance(label, str):
        return ""

    #  FIX: Replace unicode hyphens and dashes with regular hyphens
    # Handle: "–" (en-dash), "—" (em-dash), "−" (minus sign), etc.
    label = label.replace("–", "-").replace("—", "-").replace("−", "-")
    label = label.replace("&", "and").replace("+", "and")

    # Normalize unicode characters
    ascii_label = (
        unicodedata.normalize("NFKD", label).encode("ascii", "ignore").decode("ascii")
    )

    lowered = ascii_label.lower()

    #  FIX: Preserve hyphens and ampersands, but replace other special chars with spaces
    # Keep: letters, numbers, spaces, hyphens
    cleaned = re.sub(r"[^a-z0-9\s\-]", " ", lowered)

    # Normalize whitespace
    cleaned = WHITESPACE_RE.sub(" ", cleaned).strip()

    return cleaned


def clean_numeric(value: Any) -> int:
    """
    Normalize numeric values coming from Excel cells.

    IMPORTANT: This function preserves negative values exactly as they appear.
    It handles:
    - Parentheses notation: "(1,500,000)" → -1500000
    - Minus signs: "-1,500,000" → -1500000
    - Currency prefixes: "KES 1,500,000" → 1500000
    - Commas: "1,500,000" → 1500000

    The function does NOT modify or "correct" values - it only parses them.

    Raises:
        ValueNormalizationError: if the value cannot represent a number.
    """
    if value is None:
        raise ValueNormalizationError("Expected numeric value but got None.")

    if isinstance(value, bool):
        raise ValueNormalizationError("Boolean values are not valid financial amounts.")

    if isinstance(value, (int, Decimal)):
        return int(Decimal(value).to_integral_value(rounding=ROUND_HALF_UP))

    if isinstance(value, float):
        return int(Decimal(str(value)).to_integral_value(rounding=ROUND_HALF_UP))

    if isinstance(value, str):
        text = value.strip()
        if not text:
            raise ValueNormalizationError(
                "Empty string cannot be normalized to a number."
            )

        negative = False
        if text.startswith("(") and text.endswith(")"):
            negative = True
            text = text[1:-1].strip()

        text = CURRENCY_PREFIX_RE.sub("", text)
        text = text.replace("KES", "").replace("ksh", "").replace("KSH", "")
        text = text.replace(",", "")
        text = text.replace(" ", "")

        if text.startswith("-"):
            negative = True
            text = text[1:]

        sanitized = NUMERIC_RE.sub("", text)
        if sanitized.startswith("-"):
            negative = True
            sanitized = sanitized[1:]
        if "-" in sanitized:
            raise ValueNormalizationError(f"Invalid numeric token: '{value}'.")

        if sanitized.count(".") > 1 or not sanitized:
            raise ValueNormalizationError(f"Invalid numeric token: '{value}'.")

        try:
            number = Decimal(sanitized)
        except InvalidOperation as exc:
            raise ValueNormalizationError(f"Could not parse number '{value}'.") from exc

        if negative:
            number *= -1

        return int(number.to_integral_value(rounding=ROUND_HALF_UP))

    raise ValueNormalizationError(f"Unsupported value type: {type(value).__name__}.")


@dataclass(frozen=True)
class LabelMatch:
    """Represents a label match result."""

    field: str
    score: float


class LabelMatcher:
    """
    Fuzzy label matcher for financial statement fields.

     FIX: Improved fuzzy matching with better threshold handling and partial matches.
    """

    def __init__(
        self, labels: Mapping[str, Sequence[str]], threshold: float = 0.75
    ) -> None:
        #  FIX: Lower threshold to 0.75 to catch more variations
        self.threshold = threshold
        self.labels = {
            field: [normalize_label_text(label) for label in variations]
            for field, variations in labels.items()
        }

    def match(self, label: str) -> Optional[str]:
        """
        Match a label to a field using fuzzy matching.

         FIX: Improved matching with:
        - Exact match first (fastest)
        - Partial word matching (e.g., "sales revenue" matches "revenue")
        - Fuzzy string matching with SequenceMatcher
        """
        normalized = normalize_label_text(label)
        if not normalized:
            return None

        best_match: Optional[LabelMatch] = None

        # Split normalized label into words for partial matching
        label_words = set(normalized.split())

        for field, variations in self.labels.items():
            for variation in variations:
                #  FIX: Exact match (highest priority)
                if normalized == variation:
                    return field

                #  FIX: Check if all words in variation are in label (partial match)
                variation_words = set(variation.split())
                if variation_words and label_words:
                    # If variation words are subset of label words, it's a good match
                    if variation_words.issubset(label_words) or label_words.issubset(
                        variation_words
                    ):
                        # Calculate score based on word overlap
                        overlap = len(variation_words & label_words) / max(
                            len(variation_words), len(label_words)
                        )
                        if overlap >= 0.7:  # 70% word overlap
                            score = 0.9  # High score for word-based match
                            if best_match is None or score > best_match.score:
                                best_match = LabelMatch(field=field, score=score)
                            continue

                #  FIX: Fuzzy string matching (fallback)
                score = SequenceMatcher(None, normalized, variation).ratio()

                #  FIX: Boost score if key words match
                if any(
                    word in normalized for word in variation.split() if len(word) > 3
                ):
                    score = min(1.0, score + 0.1)  # Boost by 0.1

                if best_match is None or score > best_match.score:
                    best_match = LabelMatch(field=field, score=score)

        #  FIX: Return match if score meets threshold
        if best_match and best_match.score >= self.threshold:
            logger.debug(
                " Matched label '%s' to field '%s' with score %.2f",
                label,
                best_match.field,
                best_match.score,
            )
            return best_match.field

        return None


class WorksheetAccessor:
    """Provides normalized access to worksheet values, including merged cells."""

    def __init__(self, worksheet: Worksheet) -> None:
        self.worksheet = worksheet
        self.max_row = worksheet.max_row or 0
        self.max_column = worksheet.max_column or 0
        self._merged_map: Dict[Tuple[int, int], Any] = {}

        for merged_range in worksheet.merged_cells.ranges:
            anchor_value = worksheet.cell(
                merged_range.min_row, merged_range.min_col
            ).value
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    self._merged_map[(row, col)] = anchor_value

    def value(self, row: int, column: int) -> Any:
        cell_value = self.worksheet.cell(row=row, column=column).value
        if cell_value is None:
            return self._merged_map.get((row, column))
        return cell_value

    def iter_coordinates(self) -> Iterable[Tuple[int, int]]:
        for row in range(1, self.max_row + 1):
            for column in range(1, self.max_column + 1):
                yield row, column
