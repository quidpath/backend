from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from excel_extractor.intelligent import IntelligentStatementExtractor


def _build_income_statement(ws):
    ws["A1"] = "Income Statement"
    ws["A3"] = "Total Revenue"
    ws["B3"] = "12,500,000"
    ws["A4"] = "Cost of Goods Sold"
    ws["B4"] = "10,200,000"
    ws["A5"] = "Gross Profit"
    ws["B5"] = "2,300,000"
    ws["A6"] = "Operating Expenses"
    ws["B6"] = "3,800,000"
    ws["A7"] = "Operating Income"
    ws["B7"] = "-1,500,000"
    ws["A8"] = "Interest Expense"
    ws["B8"] = "950,000"
    ws["A9"] = "Taxes"
    ws["B9"] = "120,000"
    ws["A10"] = "Net Income"
    ws["B10"] = "-2,570,000"


def _build_balance_sheet(ws):
    ws["D3"] = "Total Assets"
    ws["E3"] = "5,000,000"
    ws["D4"] = "Total Liabilities"
    ws["E4"] = "3,000,000"
    ws["D5"] = "Shareholders Equity"
    ws["E5"] = "2,000,000"


def test_intelligent_extractor_reads_multiple_statements(tmp_path):
    wb = Workbook()
    ws = wb.active
    _build_income_statement(ws)
    _build_balance_sheet(ws)

    file_path = tmp_path / "financials.xlsx"
    wb.save(file_path)

    extractor = IntelligentStatementExtractor()
    result = extractor.extract(file_path)

    assert result["success"]
    statements = result["statements"]
    assert "income_statement" in statements
    assert "balance_sheet" in statements
    assert statements["income_statement"]["total_revenue"] == 12_500_000
    assert statements["balance_sheet"]["total_assets"] == 5_000_000


