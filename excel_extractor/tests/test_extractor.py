"""Pytest suite for the Excel financial extractor."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from excel_extractor.extractor import FinancialExtractor


def _build_workbook(tmp_path: Path, filename: str, fn) -> Path:
    wb = Workbook()
    ws = wb.active
    fn(wb, ws)
    file_path = tmp_path / filename
    wb.save(file_path)
    return file_path


def test_extracts_messy_formatting(tmp_path):
    def configure(_wb, ws):
        ws["A1"] = "Income Statement"
        ws["A3"] = "Revenue "
        ws["B3"] = "KES 12,500,000"
        ws["A4"] = "Cost of Goods Sold"
        ws["B4"] = " 5,000,000 "
        ws["A5"] = "Gross Profit"
        ws["B5"] = "7,500,000"
        ws["A6"] = "Operating Expenses"
        ws["B6"] = "2,000,000"
        ws["A7"] = "Operating Profit"
        ws["B7"] = "5,500,000"
        ws["A8"] = "Interest Expense"
        ws["B8"] = "500,000"
        ws["A9"] = "Taxes"
        ws["B9"] = "1,500,000"
        ws["A10"] = "Net Profit"
        ws["B10"] = "3,500,000"
        ws["A12"] = "Risk Level"
        ws["B12"] = "Low"

    file_path = _build_workbook(tmp_path, "messy.xlsx", configure)
    extractor = FinancialExtractor()
    result = extractor.extract(file_path)

    assert result["status"] == "success"
    assert result["total_revenue"] == 12_500_000
    assert result["cogs"] == 5_000_000
    assert result["risk_level"] == "Low"


def test_handles_merged_cells(tmp_path):
    def configure(_wb, ws):
        ws.merge_cells("A2:A3")
        ws["A2"] = "Total Revenue"
        ws["B2"] = 1_000_000
        ws["A4"] = "COGS"
        ws.merge_cells("B4:C4")
        ws["B4"] = 400_000
        ws["A5"] = "Gross Profit"
        ws["B5"] = 600_000
        ws["A6"] = "Operating Expenses"
        ws["B6"] = 200_000
        ws["A7"] = "Operating Income"
        ws["B7"] = 400_000
        ws["A8"] = "Interest"
        ws["B8"] = 100_000
        ws["A9"] = "Taxes"
        ws["B9"] = 50_000
        ws["A10"] = "Net Income"
        ws["B10"] = 250_000

    file_path = _build_workbook(tmp_path, "merged.xlsx", configure)
    extractor = FinancialExtractor()
    result = extractor.extract(file_path)

    assert result["status"] == "success"
    assert result["gross_profit"] == 600_000


def test_missing_field_returns_error(tmp_path):
    def configure(_wb, ws):
        ws["A1"] = "Revenue"
        ws["B1"] = 100
        ws["A2"] = "COGS"
        ws["B2"] = 40
        ws["A3"] = "Gross Profit"
        ws["B3"] = 60
        ws["A4"] = "Operating Expenses"
        ws["B4"] = 20
        ws["A5"] = "Operating Income"
        ws["B5"] = 40
        ws["A6"] = "Interest Expense"
        ws["B6"] = 5
        ws["A7"] = "Taxes"
        ws["B7"] = 10
        # Net income missing

    file_path = _build_workbook(tmp_path, "missing.xlsx", configure)
    extractor = FinancialExtractor()
    result = extractor.extract(file_path)

    assert result["status"] == "error"
    assert "Missing required financial fields" in result["message"]


def test_negative_parentheses(tmp_path):
    def configure(_wb, ws):
        ws["A1"] = "Revenue"
        ws["B1"] = "1,000"
        ws["A2"] = "Cost of Revenue"
        ws["B2"] = "(200)"
        ws["A3"] = "Gross Profit"
        ws["B3"] = "1,200"
        ws["A4"] = "Expenses"
        ws["B4"] = "(100)"
        ws["A5"] = "EBIT"
        ws["B5"] = "1,300"
        ws["A6"] = "Interest Expense"
        ws["B6"] = "(50)"
        ws["A7"] = "Taxes"
        ws["B7"] = "(150)"
        ws["A8"] = "Net Profit"
        ws["B8"] = "1,500"

    file_path = _build_workbook(tmp_path, "negative.xlsx", configure)
    extractor = FinancialExtractor()
    result = extractor.extract(file_path)

    assert result["status"] == "success"
    assert result["cogs"] == -200


def test_multiple_tables_in_sheet(tmp_path):
    def configure(_wb, ws):
        ws["A2"] = "Random Header"
        ws["C2"] = "Another table"

        ws["A4"] = "Sales"
        ws["B4"] = 2000
        ws["A5"] = "COGS"
        ws["B5"] = 800
        ws["A6"] = "Gross Profit"
        ws["B6"] = 1200
        ws["A7"] = "Operating Expenses"
        ws["B7"] = 200
        ws["A8"] = "Operating Profit"
        ws["B8"] = 1000
        ws["A9"] = "Interest Expense"
        ws["B9"] = 100
        ws["A10"] = "Taxes"
        ws["B10"] = 300
        ws["A11"] = "Net Income"
        ws["B11"] = 600

        ws["D4"] = "Risk"
        ws["E4"] = "Medium"

    file_path = _build_workbook(tmp_path, "tables.xlsx", configure)
    extractor = FinancialExtractor()
    result = extractor.extract(file_path)

    assert result["status"] == "success"
    assert result["net_income"] == 600
    assert result["risk_level"] == "Medium"


def test_corrupted_number_returns_error(tmp_path):
    def configure(_wb, ws):
        ws["A1"] = "Revenue"
        ws["B1"] = "abc"

    file_path = _build_workbook(tmp_path, "corrupted.xlsx", configure)
    extractor = FinancialExtractor()
    result = extractor.extract(file_path)

    assert result["status"] == "error"
    assert "Unable to find numeric value" in result["message"] or "Invalid numeric token" in result["message"]


def test_decimal_and_currency_prefix(tmp_path):
    def configure(_wb, ws):
        ws["A1"] = "Turnover"
        ws["B1"] = "Ksh 1,000.6"
        ws["A2"] = "Cost of Goods Sold"
        ws["B2"] = "Ksh 249.6"
        ws["A3"] = "Gross Profit"
        ws["B3"] = "751.0"
        ws["A4"] = "Operating Expenses"
        ws["B4"] = "200.4"
        ws["A5"] = "Operating Income"
        ws["B5"] = "551.4"
        ws["A6"] = "Interest Expense"
        ws["B6"] = "50.4"
        ws["A7"] = "Taxes"
        ws["B7"] = "200.4"
        ws["A8"] = "Net Income"
        ws["B8"] = "301.4"

    file_path = _build_workbook(tmp_path, "currency.xlsx", configure)
    extractor = FinancialExtractor()
    result = extractor.extract(file_path)

    assert result["status"] == "success"
    assert result["total_revenue"] == 1001
    assert result["net_income"] == 301

