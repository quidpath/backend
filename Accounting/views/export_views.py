"""
export_views.py — Professional Excel, CSV export for all entities.
Uses openpyxl with full styling: branded header, alternating rows, totals row.
"""
import csv
import io
from datetime import date, datetime
from decimal import Decimal

from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt

from quidpath_backend.core.utils.json_response import ResponseProvider
from quidpath_backend.core.utils.Logbase import TransactionLogBase
from quidpath_backend.core.utils.registry import ServiceRegistry
from quidpath_backend.core.utils.request_parser import get_clean_data


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _safe_date(value):
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        try:
            return date.fromisoformat(value)
        except ValueError:
            return None
    return None


def _get_corporate(registry, user):
    """Returns (corporate_id, corporate_dict, error_response)."""
    user_id = user.get("id") if isinstance(user, dict) else getattr(user, "id", None)
    if not user_id:
        return None, None, ResponseProvider(message="User ID not found", code=400).bad_request()
    corp_users = registry.database("CorporateUser", "filter",
                                   data={"customuser_ptr_id": user_id, "is_active": True})
    if not corp_users:
        return None, None, ResponseProvider(message="No corporate association", code=400).bad_request()
    corporate_id = corp_users[0]["corporate_id"]
    corporates = registry.database("Corporate", "filter", data={"id": corporate_id})
    corporate = corporates[0] if corporates else {}
    return corporate_id, corporate, None


def _fmt(value):
    """Format a numeric value safely."""
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _excel_response(filename):
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _csv_response(filename):
    resp = HttpResponse(content_type="text/csv")
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    return resp


def _write_excel(rows, headers, sheet_title, corporate_name="", report_title="",
                 period="", numeric_cols=None, total_cols=None):
    """
    Build a professionally styled xlsx workbook.
    - Branded header band with company name + report title
    - Bold, green column headers
    - Alternating row shading
    - Right-aligned numeric columns
    - Totals row in bold green
    - Auto-width columns
    """
    import openpyxl
    from openpyxl.styles import (Alignment, Border, Font, GradientFill,
                                  PatternFill, Side)
    from openpyxl.utils import get_column_letter

    BRAND_GREEN = "1B5E20"
    HEADER_GREEN = "2E7D32"
    LIGHT_GREEN = "E8F5E9"
    ALT_ROW = "F5F5F5"
    TOTAL_BG = "C8E6C9"
    BORDER_COLOR = "BDBDBD"

    thin = Side(style="thin", color=BORDER_COLOR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]  # Excel sheet name limit

    num_cols = len(headers)
    numeric_cols = numeric_cols or []
    total_cols = total_cols or []

    # ── Row 1: Company name banner ─────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    banner = ws.cell(row=1, column=1, value=corporate_name or "QuidPath ERP")
    banner.font = Font(bold=True, size=16, color="FFFFFF", name="Calibri")
    banner.fill = PatternFill("solid", fgColor=BRAND_GREEN)
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── Row 2: Report title ────────────────────────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    title_cell = ws.cell(row=2, column=1, value=report_title or sheet_title)
    title_cell.font = Font(bold=True, size=13, color="FFFFFF", name="Calibri")
    title_cell.fill = PatternFill("solid", fgColor=HEADER_GREEN)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    # ── Row 3: Period / metadata ───────────────────────────────────────────────
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=num_cols)
    meta_text = f"Period: {period}" if period else f"Generated: {date.today().strftime('%d %B %Y')}"
    meta_cell = ws.cell(row=3, column=1, value=meta_text)
    meta_cell.font = Font(italic=True, size=10, color="555555", name="Calibri")
    meta_cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    meta_cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[3].height = 18

    # ── Row 4: blank spacer ────────────────────────────────────────────────────
    ws.row_dimensions[4].height = 6

    # ── Row 5: Column headers ──────────────────────────────────────────────────
    header_row = 5
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=HEADER_GREEN)
        cell.alignment = Alignment(
            horizontal="right" if (col_idx - 1) in numeric_cols else "center",
            vertical="center"
        )
        cell.border = border
    ws.row_dimensions[header_row].height = 20

    # ── Data rows ──────────────────────────────────────────────────────────────
    totals = {c: 0.0 for c in total_cols}

    for row_idx, row in enumerate(rows, header_row + 1):
        fill = PatternFill("solid", fgColor=ALT_ROW) if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(size=10, name="Calibri")
            cell.border = border
            if fill:
                cell.fill = fill
            is_num = (col_idx - 1) in numeric_cols
            if is_num:
                cell.alignment = Alignment(horizontal="right")
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
                    if (col_idx - 1) in total_cols:
                        totals[col_idx - 1] = totals.get(col_idx - 1, 0) + value
            else:
                cell.alignment = Alignment(horizontal="left")

    # ── Totals row ─────────────────────────────────────────────────────────────
    if total_cols and rows:
        totals_row = header_row + len(rows) + 1
        ws.row_dimensions[totals_row].height = 20
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=totals_row, column=col_idx)
            cell.fill = PatternFill("solid", fgColor=TOTAL_BG)
            cell.font = Font(bold=True, size=10, name="Calibri")
            cell.border = border
            if col_idx == 1:
                cell.value = "TOTAL"
                cell.alignment = Alignment(horizontal="left")
            elif (col_idx - 1) in total_cols:
                cell.value = totals.get(col_idx - 1, 0)
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")

    # ── Auto-width ─────────────────────────────────────────────────────────────
    for col_idx, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header))
        for row in rows:
            if col_idx - 1 < len(row):
                max_len = max(max_len, len(str(row[col_idx - 1] or "")))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 45)

    # ── Freeze panes below header ──────────────────────────────────────────────
    ws.freeze_panes = f"A{header_row + 1}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─── Invoice Export ────────────────────────────────────────────────────────────

INVOICE_HEADERS = ["Invoice #", "Customer", "Date", "Due Date", "Status",
                   "Sub Total", "Tax", "Total", "Salesperson"]
INVOICE_NUMERIC = [5, 6, 7]
INVOICE_TOTAL = [5, 6, 7]


@csrf_exempt
def export_invoices(request):
    data, metadata = get_clean_data(request)
    user = metadata.get("user")
    if not user:
        return ResponseProvider(message="User not authenticated", code=401).unauthorized()

    registry = ServiceRegistry()
    corporate_id, corporate, err = _get_corporate(registry, user)
    if err:
        return err

    fmt = (data.get("format") or request.GET.get("format", "excel")).lower()
    start_date = _safe_date(data.get("start_date") or request.GET.get("start_date"))
    end_date = _safe_date(data.get("end_date") or request.GET.get("end_date"))

    invoices = registry.database("Invoices", "filter", data={"corporate_id": corporate_id})

    # Filter by date
    if start_date or end_date:
        invoices = [inv for inv in invoices if _in_range(_safe_date(inv.get("date")), start_date, end_date)]

    # Build customer lookup
    customer_ids = list({inv.get("customer_id") for inv in invoices if inv.get("customer_id")})
    customers = registry.database("Customer", "filter", data={"id__in": customer_ids}) if customer_ids else []
    cust_map = {str(c["id"]): c.get("name", "") for c in customers}

    # Build salesperson lookup
    sp_ids = list({inv.get("salesperson_id") for inv in invoices if inv.get("salesperson_id")})
    sps = registry.database("CorporateUser", "filter", data={"id__in": sp_ids}) if sp_ids else []
    sp_map = {str(s["id"]): s.get("username", "") for s in sps}

    rows = [[
        inv.get("number", ""),
        cust_map.get(str(inv.get("customer_id", "")), ""),
        str(inv.get("date", "")),
        str(inv.get("due_date", "")),
        inv.get("status", ""),
        _fmt(inv.get("sub_total", 0)),
        _fmt(inv.get("tax_total", 0)),
        _fmt(inv.get("total", 0)),
        sp_map.get(str(inv.get("salesperson_id", "")), ""),
    ] for inv in invoices]

    period = _period_str(start_date, end_date)
    filename_base = f"invoices_{date.today()}"

    if fmt == "csv":
        resp = _csv_response(f"{filename_base}.csv")
        writer = csv.writer(resp)
        writer.writerow(INVOICE_HEADERS)
        writer.writerows(rows)
        return resp

    content = _write_excel(rows, INVOICE_HEADERS, "Invoices",
                           corporate_name=corporate.get("name", ""),
                           report_title="Invoices Report",
                           period=period,
                           numeric_cols=INVOICE_NUMERIC,
                           total_cols=INVOICE_TOTAL)
    resp = _excel_response(f"{filename_base}.xlsx")
    resp.write(content)
    return resp


def _in_range(d, start, end):
    if d is None:
        return True
    if start and d < start:
        return False
    if end and d > end:
        return False
    return True


def _period_str(start, end):
    if start and end:
        return f"{start.strftime('%d %b %Y')} – {end.strftime('%d %b %Y')}"
    if start:
        return f"From {start.strftime('%d %b %Y')}"
    if end:
        return f"Up to {end.strftime('%d %b %Y')}"
    return f"All time (as of {date.today().strftime('%d %b %Y')})"


# ─── Vendor Bills Export ───────────────────────────────────────────────────────

BILL_HEADERS = ["Bill #", "Vendor", "Date", "Due Date", "Status",
                "Sub Total", "Tax", "Total"]
BILL_NUMERIC = [5, 6, 7]
BILL_TOTAL = [5, 6, 7]


@csrf_exempt
def export_vendor_bills(request):
    data, metadata = get_clean_data(request)
    user = metadata.get("user")
    if not user:
        return ResponseProvider(message="User not authenticated", code=401).unauthorized()

    registry = ServiceRegistry()
    corporate_id, corporate, err = _get_corporate(registry, user)
    if err:
        return err

    fmt = (data.get("format") or request.GET.get("format", "excel")).lower()
    start_date = _safe_date(data.get("start_date") or request.GET.get("start_date"))
    end_date = _safe_date(data.get("end_date") or request.GET.get("end_date"))

    bills = registry.database("VendorBill", "filter", data={"corporate_id": corporate_id})
    if start_date or end_date:
        bills = [b for b in bills if _in_range(_safe_date(b.get("date")), start_date, end_date)]

    vendor_ids = list({b.get("vendor_id") for b in bills if b.get("vendor_id")})
    vendors = registry.database("Vendor", "filter", data={"id__in": vendor_ids}) if vendor_ids else []
    vendor_map = {str(v["id"]): v.get("name", "") for v in vendors}

    rows = [[
        b.get("number", ""),
        vendor_map.get(str(b.get("vendor_id", "")), ""),
        str(b.get("date", "")),
        str(b.get("due_date", "")),
        b.get("status", ""),
        _fmt(b.get("sub_total", 0)),
        _fmt(b.get("tax_total", 0)),
        _fmt(b.get("total", 0)),
    ] for b in bills]

    period = _period_str(start_date, end_date)
    filename_base = f"vendor_bills_{date.today()}"

    if fmt == "csv":
        resp = _csv_response(f"{filename_base}.csv")
        writer = csv.writer(resp)
        writer.writerow(BILL_HEADERS)
        writer.writerows(rows)
        return resp

    content = _write_excel(rows, BILL_HEADERS, "Vendor Bills",
                           corporate_name=corporate.get("name", ""),
                           report_title="Vendor Bills Report",
                           period=period,
                           numeric_cols=BILL_NUMERIC,
                           total_cols=BILL_TOTAL)
    resp = _excel_response(f"{filename_base}.xlsx")
    resp.write(content)
    return resp


# ─── Expenses Export ───────────────────────────────────────────────────────────

EXPENSE_HEADERS = ["Date", "Vendor", "Category", "Description",
                   "Amount", "Tax", "Total", "Status", "Payment Method", "Reference"]
EXPENSE_NUMERIC = [4, 5, 6]
EXPENSE_TOTAL = [4, 5, 6]


@csrf_exempt
def export_expenses(request):
    data, metadata = get_clean_data(request)
    user = metadata.get("user")
    if not user:
        return ResponseProvider(message="User not authenticated", code=401).unauthorized()

    registry = ServiceRegistry()
    corporate_id, corporate, err = _get_corporate(registry, user)
    if err:
        return err

    fmt = (data.get("format") or request.GET.get("format", "excel")).lower()
    start_date = _safe_date(data.get("start_date") or request.GET.get("start_date"))
    end_date = _safe_date(data.get("end_date") or request.GET.get("end_date"))

    expenses = registry.database("Expense", "filter", data={"corporate_id": corporate_id})
    if start_date or end_date:
        expenses = [e for e in expenses if _in_range(_safe_date(e.get("date")), start_date, end_date)]

    vendor_ids = list({e.get("vendor_id") for e in expenses if e.get("vendor_id")})
    vendors = registry.database("Vendor", "filter", data={"id__in": vendor_ids}) if vendor_ids else []
    vendor_map = {str(v["id"]): v.get("name", "") for v in vendors}

    rows = [[
        str(e.get("date", "")),
        vendor_map.get(str(e.get("vendor_id", "")), ""),
        e.get("category", ""),
        e.get("description", ""),
        _fmt(e.get("amount", 0)),
        _fmt(e.get("tax_amount", 0) or 0),
        _fmt(e.get("amount", 0)) + _fmt(e.get("tax_amount", 0) or 0),
        e.get("status", ""),
        e.get("payment_method", ""),
        e.get("reference", ""),
    ] for e in expenses]

    period = _period_str(start_date, end_date)
    filename_base = f"expenses_{date.today()}"

    if fmt == "csv":
        resp = _csv_response(f"{filename_base}.csv")
        writer = csv.writer(resp)
        writer.writerow(EXPENSE_HEADERS)
        writer.writerows(rows)
        return resp

    content = _write_excel(rows, EXPENSE_HEADERS, "Expenses",
                           corporate_name=corporate.get("name", ""),
                           report_title="Expenses Report",
                           period=period,
                           numeric_cols=EXPENSE_NUMERIC,
                           total_cols=EXPENSE_TOTAL)
    resp = _excel_response(f"{filename_base}.xlsx")
    resp.write(content)
    return resp


# ─── Quotations Export ─────────────────────────────────────────────────────────

QUOTE_HEADERS = ["Quote #", "Customer", "Date", "Valid Until", "Status",
                 "Sub Total", "Tax", "Total"]
QUOTE_NUMERIC = [5, 6, 7]
QUOTE_TOTAL = [5, 6, 7]


@csrf_exempt
def export_quotations(request):
    data, metadata = get_clean_data(request)
    user = metadata.get("user")
    if not user:
        return ResponseProvider(message="User not authenticated", code=401).unauthorized()

    registry = ServiceRegistry()
    corporate_id, corporate, err = _get_corporate(registry, user)
    if err:
        return err

    fmt = (data.get("format") or request.GET.get("format", "excel")).lower()
    start_date = _safe_date(data.get("start_date") or request.GET.get("start_date"))
    end_date = _safe_date(data.get("end_date") or request.GET.get("end_date"))

    quotes = registry.database("Quotation", "filter", data={"corporate_id": corporate_id})
    if start_date or end_date:
        quotes = [q for q in quotes if _in_range(_safe_date(q.get("date")), start_date, end_date)]

    customer_ids = list({q.get("customer_id") for q in quotes if q.get("customer_id")})
    customers = registry.database("Customer", "filter", data={"id__in": customer_ids}) if customer_ids else []
    cust_map = {str(c["id"]): c.get("name", "") for c in customers}

    # Quotation doesn't have sub_total/tax_total/total stored — compute from lines
    rows = []
    for q in quotes:
        lines = registry.database("QuotationLine", "filter", data={"quotation_id": q["id"]})
        sub = sum(_fmt(l.get("sub_total", 0)) for l in lines)
        tax = sum(_fmt(l.get("tax_amount", 0)) for l in lines)
        total = sum(_fmt(l.get("total", 0)) for l in lines)
        rows.append([
            q.get("number", ""),
            cust_map.get(str(q.get("customer_id", "")), ""),
            str(q.get("date", "")),
            str(q.get("valid_until", "")),
            q.get("status", ""),
            sub, tax, total,
        ])

    period = _period_str(start_date, end_date)
    filename_base = f"quotations_{date.today()}"

    if fmt == "csv":
        resp = _csv_response(f"{filename_base}.csv")
        writer = csv.writer(resp)
        writer.writerow(QUOTE_HEADERS)
        writer.writerows(rows)
        return resp

    content = _write_excel(rows, QUOTE_HEADERS, "Quotations",
                           corporate_name=corporate.get("name", ""),
                           report_title="Quotations Report",
                           period=period,
                           numeric_cols=QUOTE_NUMERIC,
                           total_cols=QUOTE_TOTAL)
    resp = _excel_response(f"{filename_base}.xlsx")
    resp.write(content)
    return resp


# ─── Purchase Orders Export ────────────────────────────────────────────────────

PO_HEADERS = ["PO #", "Vendor", "Date", "Expected Delivery", "Status",
              "Sub Total", "Tax", "Total"]
PO_NUMERIC = [5, 6, 7]
PO_TOTAL = [5, 6, 7]


@csrf_exempt
def export_purchase_orders(request):
    data, metadata = get_clean_data(request)
    user = metadata.get("user")
    if not user:
        return ResponseProvider(message="User not authenticated", code=401).unauthorized()

    registry = ServiceRegistry()
    corporate_id, corporate, err = _get_corporate(registry, user)
    if err:
        return err

    fmt = (data.get("format") or request.GET.get("format", "excel")).lower()
    start_date = _safe_date(data.get("start_date") or request.GET.get("start_date"))
    end_date = _safe_date(data.get("end_date") or request.GET.get("end_date"))

    pos = registry.database("PurchaseOrder", "filter", data={"corporate_id": corporate_id})
    if start_date or end_date:
        pos = [p for p in pos if _in_range(_safe_date(p.get("date")), start_date, end_date)]

    vendor_ids = list({p.get("vendor_id") for p in pos if p.get("vendor_id")})
    vendors = registry.database("Vendor", "filter", data={"id__in": vendor_ids}) if vendor_ids else []
    vendor_map = {str(v["id"]): v.get("name", "") for v in vendors}

    rows = []
    for p in pos:
        lines = registry.database("PurchaseOrderLine", "filter", data={"purchase_order_id": p["id"]})
        sub = sum(_fmt(l.get("sub_total", 0)) for l in lines)
        tax = sum(_fmt(l.get("tax_amount", 0)) for l in lines)
        total = sum(_fmt(l.get("total", 0)) for l in lines)
        rows.append([
            p.get("number", ""),
            vendor_map.get(str(p.get("vendor_id", "")), ""),
            str(p.get("date", "")),
            str(p.get("expected_delivery", "")),
            p.get("status", ""),
            sub, tax, total,
        ])

    period = _period_str(start_date, end_date)
    filename_base = f"purchase_orders_{date.today()}"

    if fmt == "csv":
        resp = _csv_response(f"{filename_base}.csv")
        writer = csv.writer(resp)
        writer.writerow(PO_HEADERS)
        writer.writerows(rows)
        return resp

    content = _write_excel(rows, PO_HEADERS, "Purchase Orders",
                           corporate_name=corporate.get("name", ""),
                           report_title="Purchase Orders Report",
                           period=period,
                           numeric_cols=PO_NUMERIC,
                           total_cols=PO_TOTAL)
    resp = _excel_response(f"{filename_base}.xlsx")
    resp.write(content)
    return resp


# ─── Journal Entries Export ────────────────────────────────────────────────────

JOURNAL_HEADERS = ["Date", "Reference", "Description", "Account Code",
                   "Account Name", "Debit", "Credit", "Balance", "Status"]
JOURNAL_NUMERIC = [5, 6, 7]
JOURNAL_TOTAL = [5, 6]


@csrf_exempt
def export_journal_entries(request):
    data, metadata = get_clean_data(request)
    user = metadata.get("user")
    if not user:
        return ResponseProvider(message="User not authenticated", code=401).unauthorized()

    registry = ServiceRegistry()
    corporate_id, corporate, err = _get_corporate(registry, user)
    if err:
        return err

    fmt = (data.get("format") or request.GET.get("format", "excel")).lower()
    start_date = _safe_date(data.get("start_date") or request.GET.get("start_date"))
    end_date = _safe_date(data.get("end_date") or request.GET.get("end_date"))

    journal_entries = registry.database("JournalEntry", "filter",
                                        data={"corporate_id": corporate_id})
    if start_date or end_date:
        journal_entries = [je for je in journal_entries
                           if _in_range(_safe_date(je.get("date")), start_date, end_date)]

    je_ids = {je["id"] for je in journal_entries}
    je_map = {je["id"]: je for je in journal_entries}

    all_lines = registry.database("JournalEntryLine", "filter", data={})
    lines = [l for l in all_lines if l["journal_entry_id"] in je_ids]

    account_ids = list({l["account_id"] for l in lines})
    accounts = registry.database("Account", "filter",
                                 data={"id__in": account_ids, "corporate_id": corporate_id}) if account_ids else []
    acc_map = {a["id"]: a for a in accounts}

    rows = []
    running = {}
    for line in sorted(lines, key=lambda l: str(je_map.get(l["journal_entry_id"], {}).get("date", ""))):
        je = je_map.get(line["journal_entry_id"], {})
        acc = acc_map.get(line["account_id"], {})
        debit = _fmt(line.get("debit", 0))
        credit = _fmt(line.get("credit", 0))
        acc_id = line["account_id"]
        running[acc_id] = running.get(acc_id, 0.0) + debit - credit
        rows.append([
            str(je.get("date", "")),
            je.get("reference", ""),
            line.get("description", "") or je.get("description", ""),
            acc.get("code", ""),
            acc.get("name", ""),
            debit,
            credit,
            round(running[acc_id], 2),
            "Posted" if je.get("is_posted") else "Draft",
        ])

    period = _period_str(start_date, end_date)
    filename_base = f"journal_entries_{date.today()}"

    if fmt == "csv":
        resp = _csv_response(f"{filename_base}.csv")
        writer = csv.writer(resp)
        writer.writerow(JOURNAL_HEADERS)
        writer.writerows(rows)
        return resp

    content = _write_excel(rows, JOURNAL_HEADERS, "Journal Entries",
                           corporate_name=corporate.get("name", ""),
                           report_title="General Journal",
                           period=period,
                           numeric_cols=JOURNAL_NUMERIC,
                           total_cols=JOURNAL_TOTAL)
    resp = _excel_response(f"{filename_base}.xlsx")
    resp.write(content)
    return resp


# ─── Financial Report Export ───────────────────────────────────────────────────

@csrf_exempt
def export_financial_report(request):
    """
    Export a stored FinancialReport by report_id as styled Excel or CSV.
    Params: report_id, format (excel|csv)
    """
    data, metadata = get_clean_data(request)
    user = metadata.get("user")
    if not user:
        return ResponseProvider(message="User not authenticated", code=401).unauthorized()

    registry = ServiceRegistry()
    corporate_id, corporate, err = _get_corporate(registry, user)
    if err:
        return err

    report_id = data.get("report_id") or request.GET.get("report_id")
    fmt = (data.get("format") or request.GET.get("format", "excel")).lower()

    if not report_id:
        return ResponseProvider(message="report_id is required", code=400).bad_request()

    reports = registry.database("FinancialReport", "filter",
                                data={"id": report_id, "corporate_id": corporate_id})
    if not reports:
        return ResponseProvider(message="Report not found", code=404).bad_request()

    report = reports[0]
    report_type = report["report_type"]
    report_data = report["data"]
    start_date = report.get("start_date")
    end_date = report.get("end_date", date.today())
    period = _period_str(
        _safe_date(str(start_date)) if start_date else None,
        _safe_date(str(end_date)) if end_date else None
    )

    headers = ["Section", "Description", "Amount"]
    rows = []

    if report_type in ("PROFIT_LOSS", "INCOME_STATEMENT"):
        rows.append(["REVENUES", "", ""])
        rev = report_data.get("revenues", {})
        rev_items = rev.get("subtypes", rev) if isinstance(rev, dict) else {}
        for desc, amt in rev_items.items():
            rows.append(["", desc, _fmt(amt)])
        rows.append(["", "Total Revenues", _fmt(report_data.get("total_revenues", 0) or
                                                 report_data.get("revenues", {}).get("total", 0))])
        rows.append(["", "", ""])
        rows.append(["EXPENSES", "", ""])
        exp = report_data.get("expenses", {})
        exp_items = exp.get("subtypes", exp) if isinstance(exp, dict) else {}
        for desc, amt in exp_items.items():
            rows.append(["", desc, _fmt(amt)])
        rows.append(["", "Total Expenses", _fmt(report_data.get("total_expenses", 0) or
                                                  report_data.get("expenses", {}).get("total", 0))])
        rows.append(["", "", ""])
        rows.append(["NET PROFIT / (LOSS)", "", _fmt(report_data.get("net_profit", 0) or
                                                       report_data.get("net_income", 0))])
        title = "Profit & Loss Statement"

    elif report_type == "BALANCE_SHEET":
        for section, label in [("assets", "ASSETS"), ("liabilities", "LIABILITIES"), ("equity", "EQUITY")]:
            rows.append([label, "", ""])
            sec = report_data.get(section, {})
            items = sec.get("subtypes", {}) if isinstance(sec, dict) else {}
            for desc, amt in items.items():
                rows.append(["", desc, _fmt(amt)])
            rows.append(["", f"Total {label.title()}", _fmt(sec.get("total", 0) if isinstance(sec, dict) else 0)])
            rows.append(["", "", ""])
        title = "Balance Sheet"

    elif report_type == "CASH_FLOW":
        for key, label in [
            ("operating_cash_flow", "Operating Activities"),
            ("investing_cash_flow", "Investing Activities"),
            ("financing_cash_flow", "Financing Activities"),
            ("net_change_in_cash", "Net Change in Cash"),
            ("net_cash_change", "Net Cash Change"),
            ("beginning_cash", "Beginning Cash Balance"),
            ("ending_cash", "Ending Cash Balance"),
        ]:
            if key in report_data:
                rows.append(["", label, _fmt(report_data[key])])
        title = "Cash Flow Statement"
    else:
        title = report_type.replace("_", " ").title()

    filename_base = f"{report_type.lower()}_{end_date}"

    if fmt == "csv":
        resp = _csv_response(f"{filename_base}.csv")
        writer = csv.writer(resp)
        writer.writerow(headers)
        writer.writerows(rows)
        return resp

    content = _write_excel(rows, headers, title,
                           corporate_name=corporate.get("name", ""),
                           report_title=title,
                           period=period,
                           numeric_cols=[2],
                           total_cols=[])
    resp = _excel_response(f"{filename_base}.xlsx")
    resp.write(content)
    return resp
