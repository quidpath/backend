"""
import_views.py — CSV/Excel import for invoices, expenses, products, customers, vendors.
Uses pandas + openpyxl for parsing.
"""
import io
from datetime import date
from decimal import Decimal, InvalidOperation

from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt

from quidpath_backend.core.utils.json_response import ResponseProvider
from quidpath_backend.core.utils.Logbase import TransactionLogBase
from quidpath_backend.core.utils.registry import ServiceRegistry
from quidpath_backend.core.utils.request_parser import get_clean_data


def _get_corporate_id(registry, user):
    user_id = user.get("id") if isinstance(user, dict) else getattr(user, "id", None)
    if not user_id:
        return None, ResponseProvider(message="User ID not found", code=400).bad_request()
    corp_users = registry.database("CorporateUser", "filter",
                                   data={"customuser_ptr_id": user_id, "is_active": True})
    if not corp_users:
        return None, ResponseProvider(message="No corporate association", code=400).bad_request()
    return corp_users[0]["corporate_id"], None


def _read_file_to_df(file_obj, filename):
    """Parse uploaded CSV or Excel file into a pandas DataFrame."""
    import pandas as pd
    name = filename.lower()
    if name.endswith(".csv"):
        return pd.read_csv(io.BytesIO(file_obj.read()))
    elif name.endswith((".xlsx", ".xls")):
        return pd.read_excel(io.BytesIO(file_obj.read()))
    raise ValueError(f"Unsupported file type: {filename}. Use CSV or Excel.")


def _safe_str(val):
    if val is None:
        return ""
    import math
    try:
        if math.isnan(float(val)):
            return ""
    except (TypeError, ValueError):
        pass
    return str(val).strip()


def _safe_decimal(val, default=Decimal("0.00")):
    try:
        return Decimal(str(val))
    except (InvalidOperation, TypeError, ValueError):
        return default


# ─── Import Customers ─────────────────────────────────────────────────────────

@csrf_exempt
def import_customers(request):
    """
    Import customers from CSV/Excel.
    Required columns: name, email
    Optional: phone, address, city, country, tax_id
    """
    data, metadata = get_clean_data(request)
    user = metadata.get("user")
    if not user:
        return ResponseProvider(message="User not authenticated", code=401).unauthorized()

    registry = ServiceRegistry()
    corporate_id, err = _get_corporate_id(registry, user)
    if err:
        return err

    file_obj = request.FILES.get("file")
    if not file_obj:
        return ResponseProvider(message="No file uploaded", code=400).bad_request()

    try:
        df = _read_file_to_df(file_obj, file_obj.name)
        df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]

        if "name" not in df.columns:
            return ResponseProvider(message="Missing required column: name", code=400).bad_request()

        created, skipped, errors = 0, 0, []
        for idx, row in df.iterrows():
            name = _safe_str(row.get("name"))
            if not name:
                skipped += 1
                continue
            try:
                registry.database("Customer", "create", data={
                    "corporate_id": corporate_id,
                    "name": name,
                    "email": _safe_str(row.get("email")),
                    "phone": _safe_str(row.get("phone")),
                    "address": _safe_str(row.get("address")),
                    "city": _safe_str(row.get("city")),
                    "country": _safe_str(row.get("country")),
                    "tax_id": _safe_str(row.get("tax_id")),
                })
                created += 1
            except Exception as e:
                errors.append(f"Row {idx + 2}: {str(e)}")

        return ResponseProvider(
            message=f"Import complete: {created} created, {skipped} skipped",
            data={"created": created, "skipped": skipped, "errors": errors[:20]},
            code=200,
        ).success()

    except Exception as e:
        return ResponseProvider(message=f"Import failed: {str(e)}", code=400).bad_request()


# ─── Import Vendors ───────────────────────────────────────────────────────────

@csrf_exempt
def import_vendors(request):
    """
    Import vendors from CSV/Excel.
    Required columns: name
    Optional: email, phone, address, city, country, tax_id, category
    """
    data, metadata = get_clean_data(request)
    user = metadata.get("user")
    if not user:
        return ResponseProvider(message="User not authenticated", code=401).unauthorized()

    registry = ServiceRegistry()
    corporate_id, err = _get_corporate_id(registry, user)
    if err:
        return err

    file_obj = request.FILES.get("file")
    if not file_obj:
        return ResponseProvider(message="No file uploaded", code=400).bad_request()

    try:
        df = _read_file_to_df(file_obj, file_obj.name)
        df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]

        if "name" not in df.columns:
            return ResponseProvider(message="Missing required column: name", code=400).bad_request()

        created, skipped, errors = 0, 0, []
        for idx, row in df.iterrows():
            name = _safe_str(row.get("name"))
            if not name:
                skipped += 1
                continue
            try:
                registry.database("Vendor", "create", data={
                    "corporate_id": corporate_id,
                    "name": name,
                    "email": _safe_str(row.get("email")),
                    "phone": _safe_str(row.get("phone")),
                    "address": _safe_str(row.get("address")),
                    "city": _safe_str(row.get("city")),
                    "country": _safe_str(row.get("country")),
                    "tax_id": _safe_str(row.get("tax_id")),
                    "category": _safe_str(row.get("category")),
                })
                created += 1
            except Exception as e:
                errors.append(f"Row {idx + 2}: {str(e)}")

        return ResponseProvider(
            message=f"Import complete: {created} created, {skipped} skipped",
            data={"created": created, "skipped": skipped, "errors": errors[:20]},
            code=200,
        ).success()

    except Exception as e:
        return ResponseProvider(message=f"Import failed: {str(e)}", code=400).bad_request()


# ─── Import Expenses ──────────────────────────────────────────────────────────

@csrf_exempt
def import_expenses(request):
    """
    Import expenses from CSV/Excel.
    Required columns: date, category, description, amount
    Optional: vendor, payment_method, reference, tax_amount
    """
    data, metadata = get_clean_data(request)
    user = metadata.get("user")
    if not user:
        return ResponseProvider(message="User not authenticated", code=401).unauthorized()

    registry = ServiceRegistry()
    corporate_id, err = _get_corporate_id(registry, user)
    if err:
        return err

    file_obj = request.FILES.get("file")
    if not file_obj:
        return ResponseProvider(message="No file uploaded", code=400).bad_request()

    try:
        df = _read_file_to_df(file_obj, file_obj.name)
        df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]

        required = {"date", "category", "description", "amount"}
        missing = required - set(df.columns)
        if missing:
            return ResponseProvider(
                message=f"Missing required columns: {', '.join(missing)}", code=400
            ).bad_request()

        # Get default vendor if needed
        vendors = registry.database("Vendor", "filter", data={"corporate_id": corporate_id})
        vendor_map = {v["name"].lower(): v["id"] for v in vendors}

        created, skipped, errors = 0, 0, []
        for idx, row in df.iterrows():
            try:
                date_val = _safe_str(row.get("date"))
                category = _safe_str(row.get("category"))
                description = _safe_str(row.get("description"))
                amount = _safe_decimal(row.get("amount"))

                if not date_val or not category or not description or amount <= 0:
                    skipped += 1
                    continue

                vendor_name = _safe_str(row.get("vendor", ""))
                vendor_id = vendor_map.get(vendor_name.lower()) if vendor_name else None

                registry.database("Expense", "create", data={
                    "corporate_id": corporate_id,
                    "date": date_val,
                    "vendor_id": vendor_id,
                    "category": category,
                    "description": description,
                    "amount": str(amount),
                    "tax_amount": str(_safe_decimal(row.get("tax_amount", 0))),
                    "payment_method": _safe_str(row.get("payment_method")),
                    "reference": _safe_str(row.get("reference")),
                    "status": "DRAFT",
                })
                created += 1
            except Exception as e:
                errors.append(f"Row {idx + 2}: {str(e)}")

        return ResponseProvider(
            message=f"Import complete: {created} created, {skipped} skipped",
            data={"created": created, "skipped": skipped, "errors": errors[:20]},
            code=200,
        ).success()

    except Exception as e:
        return ResponseProvider(message=f"Import failed: {str(e)}", code=400).bad_request()


# ─── Import Products (Inventory) ──────────────────────────────────────────────

@csrf_exempt
def import_products(request):
    """
    Import inventory products from CSV/Excel.
    Required columns: name, sku, unit_price, cost_price
    Optional: barcode, category, description, quantity_on_hand, reorder_point, unit_of_measure
    """
    data, metadata = get_clean_data(request)
    user = metadata.get("user")
    if not user:
        return ResponseProvider(message="User not authenticated", code=401).unauthorized()

    registry = ServiceRegistry()
    corporate_id, err = _get_corporate_id(registry, user)
    if err:
        return err

    file_obj = request.FILES.get("file")
    if not file_obj:
        return ResponseProvider(message="No file uploaded", code=400).bad_request()

    try:
        df = _read_file_to_df(file_obj, file_obj.name)
        df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]

        required = {"name", "sku", "unit_price", "cost_price"}
        missing = required - set(df.columns)
        if missing:
            return ResponseProvider(
                message=f"Missing required columns: {', '.join(missing)}", code=400
            ).bad_request()

        # Get default warehouse
        warehouses = registry.database("Warehouse", "filter", data={"corporate_id": corporate_id})
        default_warehouse_id = warehouses[0]["id"] if warehouses else None

        created, skipped, errors = 0, 0, []
        for idx, row in df.iterrows():
            try:
                name = _safe_str(row.get("name"))
                sku = _safe_str(row.get("sku"))
                if not name or not sku:
                    skipped += 1
                    continue

                registry.database("InventoryItem", "create", data={
                    "corporate_id": corporate_id,
                    "warehouse_id": default_warehouse_id,
                    "name": name,
                    "sku": sku,
                    "barcode": _safe_str(row.get("barcode")),
                    "category": _safe_str(row.get("category")),
                    "description": _safe_str(row.get("description")),
                    "unit_price": str(_safe_decimal(row.get("unit_price"))),
                    "cost_price": str(_safe_decimal(row.get("cost_price"))),
                    "quantity_on_hand": int(_safe_decimal(row.get("quantity_on_hand", 0))),
                    "reorder_point": int(_safe_decimal(row.get("reorder_point", 0))),
                    "unit_of_measure": _safe_str(row.get("unit_of_measure", "pcs")),
                    "is_active": True,
                })
                created += 1
            except Exception as e:
                errors.append(f"Row {idx + 2}: {str(e)}")

        return ResponseProvider(
            message=f"Import complete: {created} created, {skipped} skipped",
            data={"created": created, "skipped": skipped, "errors": errors[:20]},
            code=200,
        ).success()

    except Exception as e:
        return ResponseProvider(message=f"Import failed: {str(e)}", code=400).bad_request()


# ─── Download Import Templates ────────────────────────────────────────────────

@csrf_exempt
def download_import_template(request):
    """
    Download a blank Excel template for a given entity.
    Param: entity = customers | vendors | expenses | products
    """
    entity = request.GET.get("entity", "customers")

    templates = {
        "customers": (
            ["name", "email", "phone", "address", "city", "country", "tax_id"],
            [["Acme Corp", "acme@example.com", "+254700000000", "123 Main St", "Nairobi", "Kenya", "P051234567X"]],
        ),
        "vendors": (
            ["name", "email", "phone", "address", "city", "country", "tax_id", "category"],
            [["Supplier Ltd", "supplier@example.com", "+254700000001", "456 Vendor Ave", "Nairobi", "Kenya", "P051234568X", "Goods"]],
        ),
        "expenses": (
            ["date", "category", "description", "amount", "vendor", "payment_method", "reference", "tax_amount"],
            [["2025-01-15", "OFFICE_SUPPLIES", "Printer paper", "5000", "Supplier Ltd", "CASH", "EXP-001", "800"]],
        ),
        "products": (
            ["name", "sku", "unit_price", "cost_price", "barcode", "category", "description", "quantity_on_hand", "reorder_point", "unit_of_measure"],
            [["Widget A", "WGT-001", "1500", "900", "1234567890", "Electronics", "A quality widget", "100", "10", "pcs"]],
        ),
    }

    if entity not in templates:
        return ResponseProvider(message=f"Unknown entity: {entity}", code=400).bad_request()

    headers, sample_rows = templates[entity]

    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = entity.title()

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2E7D32")

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill

    for row_idx, row in enumerate(sample_rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{entity}_import_template.xlsx"'
    return resp
